# utils/export_utils.py
import os
from datetime import datetime
import pandas as pd
import hashlib
import time
from io import BytesIO
from django.conf import settings
from django.db import transaction
from user.models import User
from score.models import AcademicPerformance

# 导入 openpyxl 样式
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class UserExporter:
    """用户信息导出器 - 精简版"""

    @staticmethod
    def generate_filename(accounts, user_type='all'):
        """生成格式化时间戳文件名，使用下划线分隔"""
        current_time = datetime.now()
        compact_timestamp = current_time.strftime("%Y%m%d_%H%M%S_") + f"{current_time.microsecond // 1000:03d}"

        if user_type == 'all' or (len(accounts) == 1 and accounts[0] == "*"):
            filename = f"users_export_{compact_timestamp}.xlsx"
        elif user_type == 'students':
            filename = f"students_export_{compact_timestamp}.xlsx"
        elif user_type == 'teachers':
            filename = f"teachers_export_{compact_timestamp}.xlsx"
        else:
            if accounts:
                if len(accounts) == 1:
                    filename = f"user_{accounts[0]}_{compact_timestamp}.xlsx"
                else:
                    accounts_str = '_'.join(sorted(accounts))
                    if len(accounts_str) > 30:
                        accounts_hash = hashlib.md5(accounts_str.encode()).hexdigest()[:8]
                        filename = f"selected_{accounts_hash}_{compact_timestamp}.xlsx"
                    else:
                        safe_str = ''.join(c if c.isalnum() else '_' for c in accounts_str)
                        filename = f"users_{safe_str}_{compact_timestamp}.xlsx"
            else:
                filename = f"export_{compact_timestamp}.xlsx"

        print(f"📁 生成文件名: {filename}")
        return filename

    @staticmethod
    def get_export_path(filename):
        """获取导出文件完整路径"""
        export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        return os.path.join(export_dir, filename)

    @staticmethod
    def get_export_url(filename):
        """获取导出文件相对URL路径"""
        return f"/media/exports/{filename}"

    @staticmethod
    def export_users_to_excel(accounts, request_user):
        """
        导出用户信息到Excel文件，返回文件信息
        """
        print("=== 开始导出用户信息 ===")
        print(f"📋 请求用户: {request_user.school_id} ({request_user.name})")
        print(f"📋 导出账号: {accounts}")

        # 确定导出范围
        export_all = (len(accounts) == 1 and accounts[0] == "*")

        # 查询用户数据
        with transaction.atomic():
            if export_all:
                users = User.objects.all().select_related('academic_performance')
                user_type = 'all'
                print("✅ 导出范围: 所有用户")
            else:
                users = User.objects.filter(
                    school_id__in=accounts
                ).select_related('academic_performance')
                user_type = 'selected'
                print(f"✅ 导出范围: 指定账号 {len(accounts)} 个")

            user_count = users.count()
            if user_count == 0:
                raise ValueError("未找到符合条件的用户")

            print(f"✅ 查询到 {user_count} 个用户")

            # 准备数据
            data_list = UserExporter.prepare_user_data(users)

            # 生成Excel（带样式）
            excel_data = UserExporter.generate_styled_excel(data_list)

            # 生成文件名并保存
            filename = UserExporter.generate_filename(
                accounts if not export_all else [],
                user_type
            )
            filepath = UserExporter.get_export_path(filename)

            # 保存文件
            with open(filepath, 'wb') as f:
                f.write(excel_data)

            print(f"✅ Excel文件保存成功: {filename}")
            print(f"📁 文件大小: {len(excel_data)} 字节")

            # 返回文件信息
            return {
                'filename': filename,
                'filepath': filepath,
                'url': UserExporter.get_export_url(filename),
                'size': len(excel_data),
                'count': user_count,
                'export_time': int(time.time() * 1000)
            }

    @staticmethod
    def prepare_user_data(users):
        """准备用户数据（精简字段） - 增强容错性"""
        data_list = []

        for user in users:
            try:
                # 尝试获取关联的成绩信息，如果不存在则设为None
                academic = getattr(user, 'academic_performance', None)

                # 基础用户信息（精简版）
                user_data = {
                    '学号/工号': user.school_id,
                    '姓名': user.name,
                    '用户类型': UserExporter.get_user_type_display(user.user_type),
                    '学院': user.college or '',
                    '联系方式': user.contact or '',
                    '邮箱': user.email or '',
                    '创建时间': user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else '',
                }

                # 学生特定字段（精简）- 安全处理
                if user.user_type == 0:  # 学生
                    if academic:
                        # 有成绩信息
                        user_data.update({
                            '绩点(GPA)': UserExporter.format_decimal(academic.gpa),
                            '四级成绩': academic.cet4 if academic.cet4 != -1 else '未参加',
                            '六级成绩': academic.cet6 if academic.cet6 != -1 else '未参加',
                            '综合总分': UserExporter.format_decimal(academic.total_comprehensive_score),
                            '学业成绩': UserExporter.format_decimal(academic.academic_score),
                            '加权分数': UserExporter.format_decimal(academic.weighted_score),
                            '绩点排名': academic.gpa_ranking if academic.gpa_ranking else '',
                            '排名维度': academic.ranking_dimension if academic.ranking_dimension else '',
                        })
                    else:
                        # 没有成绩信息，显示空值
                        user_data.update({
                            '绩点(GPA)': '',
                            '四级成绩': '未参加',
                            '六级成绩': '未参加',
                            '综合总分': '',
                            '学业成绩': '',
                            '加权分数': '',
                            '绩点排名': '',
                            '排名维度': '未设置',
                        })
                        print(f"⚠️ 用户 {user.school_id} 缺少 AcademicPerformance 记录")

                # 教师特定字段
                elif user.user_type == 1:
                    user_data.update({
                    })

                data_list.append(user_data)

            except Exception as e:
                print(f"❌ 处理用户 {user.school_id} 数据时出错: {e}")
                # 即使出错也添加基础信息
                data_list.append({
                    '学号/工号': user.school_id,
                    '姓名': user.name,
                    '用户类型': UserExporter.get_user_type_display(user.user_type),
                    '学院': '数据错误',
                    '专业': '数据错误',
                    '联系方式': '',
                    '邮箱': '',
                    '创建时间': '',
                })

        return data_list

    @staticmethod
    def format_decimal(value):
        """格式化Decimal数值"""
        if value is None:
            return ''
        try:
            return float(value)
        except:
            return ''

    @staticmethod
    def generate_styled_excel(data_list):
        """生成带样式的Excel文件（列宽自适应 + 内容居中）"""
        if not data_list:
            raise ValueError("没有数据可以导出")

        df = pd.DataFrame(data_list)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 写入数据
            df.to_excel(writer, index=False, sheet_name='用户信息')

            # 获取worksheet
            worksheet = writer.sheets['用户信息']

            # 🔧 1. 智能调整列宽（根据最大字数适配）
            UserExporter.adjust_column_width_smart(worksheet, df)

            # 🎯 2. 设置内容居中
            UserExporter.center_all_cells(worksheet)

            # 🎨 3. 应用表格样式（可选）
            UserExporter.apply_table_styles(worksheet, df)

        return output.getvalue()

    @staticmethod
    def adjust_column_width_smart(worksheet, df):
        """智能调整列宽 - 根据最大字数适配"""
        print("🔧 开始智能调整列宽...")

        for i, column_name in enumerate(df.columns, start=1):
            column_letter = get_column_letter(i)
            column_data = df.iloc[:, i - 1]

            # 计算最大宽度
            max_width = 0

            # 1. 表头宽度
            header_width = len(str(column_name))
            if any('\u4e00' <= c <= '\u9fff' for c in str(column_name)):
                header_width *= 2  # 中文字符宽度加倍

            max_width = max(max_width, header_width)

            # 2. 数据行最大宽度
            for value in column_data:
                if pd.notna(value):
                    value_str = str(value)
                    value_width = len(value_str)

                    # 考虑中文字符宽度
                    if any('\u4e00' <= c <= '\u9fff' for c in value_str):
                        value_width = value_width * 2

                    max_width = max(max_width, value_width)

            # 3. 添加边距并设置限制
            adjusted_width = min(max(max_width + 2, 8), 50)  # 最小8，最大50

            # 4. 特定字段的特殊处理
            column_name_str = str(column_name)
            if '邮箱' in column_name_str or 'Email' in column_name_str:
                adjusted_width = min(adjusted_width, 30)
            elif '联系方式' in column_name_str or '电话' in column_name_str:
                adjusted_width = min(adjusted_width, 15)
            elif '学号' in column_name_str or '工号' in column_name_str:
                adjusted_width = min(adjusted_width, 12)

            # 设置列宽
            worksheet.column_dimensions[column_letter].width = adjusted_width

            print(f"  {column_name_str}: {adjusted_width:.1f} 字符宽度")

    @staticmethod
    def center_all_cells(worksheet):
        """设置所有单元格内容居中"""
        print("🎯 设置单元格内容居中...")

        # 创建居中对齐样式
        center_alignment = Alignment(
            horizontal='center',  # 水平居中
            vertical='center',  # 垂直居中
            wrap_text=False  # 不自动换行
        )

        # 应用到所有单元格
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = center_alignment

        # 表头加粗
        for cell in worksheet[1]:  # 第一行是表头
            cell.font = Font(bold=True, size=11)

        print(f"✅ 已完成 {worksheet.max_row} 行 × {worksheet.max_column} 列的内容居中")

    @staticmethod
    def apply_table_styles(worksheet, df):
        """应用表格样式（美化）"""
        print("🎨 应用表格样式...")

        # 创建边框样式
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # 表头背景色
        header_fill = PatternFill(
            start_color="E0E0E0",  # 浅灰色
            end_color="E0E0E0",
            fill_type="solid"
        )

        # 应用样式
        max_row = len(df) + 1
        max_col = len(df.columns)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)

                # 设置边框
                cell.border = thin_border

                # 设置表头样式
                if row == 1:
                    cell.fill = header_fill
                    cell.font = Font(bold=True, size=11, color="000000")

        print("✅ 表格样式应用完成")

    @staticmethod
    def generate_excel_old(data_list):
        """旧的生成Excel方法（保持兼容）"""
        if not data_list:
            raise ValueError("没有数据可以导出")

        df = pd.DataFrame(data_list)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='用户信息')

            worksheet = writer.sheets['用户信息']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        return output.getvalue()

    @staticmethod
    def get_user_type_display(user_type):
        """获取用户类型显示文本"""
        return {0: '学生', 1: '教师', 2: '超级管理员'}.get(user_type, '未知')

    @staticmethod
    def cleanup_old_files(max_age_hours=24, max_files=100):
        """清理旧的导出文件"""
        import glob
        import time

        export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        if not os.path.exists(export_dir):
            return

        files = glob.glob(os.path.join(export_dir, '*.xlsx'))
        files.sort(key=os.path.getmtime, reverse=True)

        # 按数量清理
        if len(files) > max_files:
            for file_to_remove in files[max_files:]:
                try:
                    os.remove(file_to_remove)
                    print(f"🗑️ 清理旧文件: {os.path.basename(file_to_remove)}")
                except:
                    pass

        # 按时间清理
        current_time = time.time()
        for filepath in files:
            try:
                file_age = current_time - os.path.getmtime(filepath)
                if file_age > max_age_hours * 3600:
                    os.remove(filepath)
                    print(f"🗑️ 清理过期文件: {os.path.basename(filepath)}")
            except:
                pass